from datetime import datetime, timedelta
import requests
import os
import re
import json_info as prop
from loguru import logger
import inquirer
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
import xml.etree.ElementTree as ET
import variables as var
import argparse
import excel as ex


def main():
    try:
        global config
        config = prop.get_config()
        start_date, end_date, restaurants_list, has_args = parse_args()
        if has_args == False:
            restaurants = []
            restaurants = get_restaurants(restaurants)
            data = questions(restaurants)
            restaurants_list, start_date, end_date = data_processing(data)
        username, password = prop.get_info_api(config)
        session = requests.Session()
        headers = {"Accept-Language": "uk,en-US;q=0.7,en;q=0.3"}
        delta = end_date - start_date
        logger.info(f"Буде скачано звіти за {delta.days+1} днів")
        for restaurant in restaurants_list:
            logger.info(f"Скачую звіти для {restaurant}")
            port, preset_id, url = set_variables(restaurant)
            auth(session, port, username, password, url)
            for day in range(delta.days+1):
                date = (start_date + timedelta(days=day)).strftime("%d.%m.%Y")
                response = generate_reports(
                    session, date, preset_id, headers, port, url)
                if len(response.text) < 1000:
                    logger.warning(
                        f"Звіт за {date} порожній або щось пішло не так. Перевір, будь ласка.")
                else:
                    create_excel_report(restaurant, date, response.text)
    except Exception as e:
        logger.error(
            "Сталася помилка. Якщо вона не описана вище, то ми її не планували :(")
        logger.error(e)
    finally:
        input("Натисни Enter для виходу")


def parse_args():
    parser = argparse.ArgumentParser()
    parser.add_argument('-sd', '--start_date')
    parser.add_argument('-ed', '--end_date')
    parser.add_argument('-r', '--restaurants', type=str, nargs='+')

    args = parser.parse_args()
    restaurants = args.restaurants
    start_date = args.start_date
    end_date = args.end_date

    if start_date == None and end_date == None and restaurants == None:
        has_args = False
    elif start_date == None or end_date == None or restaurants == None:
        logger.error("Передано частину аргументів, щось не так")
        raise Exception()
    else:
        start_date = datetime.strptime(args.start_date, '%d.%m.%Y')
        end_date = datetime.strptime(args.end_date, '%d.%m.%Y')
        has_args = True

    return start_date, end_date, restaurants, has_args


def get_restaurants(restaurants):
    try:
        for entity in config["legal_entities"]:
            if isinstance(config["legal_entities"][entity]["name"], list):
                restaurants.extend(config["legal_entities"][entity]["name"])
            else:
                restaurants.append(config["legal_entities"][entity]["name"])
        return restaurants
    except Exception as e:
        logger.error(
            f"Схоже, що в конфігураційному файлі немає полів {restaurants}. Будь ласка, перевір!")
        logger.error(e)


def set_variables(restaurant):
    for entity in config["legal_entities"]:
        if restaurant in config["legal_entities"][entity]["name"]:
            port = config["legal_entities"][entity]["port"]
            url = config["legal_entities"][entity]["url"]
            if isinstance(config["legal_entities"][entity]["name"], list):
                position = config["legal_entities"][entity]["name"].index(
                    restaurant)
                preset_id = config["legal_entities"][entity]["preset_id"][position]
            else:
                preset_id = config["legal_entities"][entity]["preset_id"]
    return port, preset_id, url


def auth(session, port, username, password, url):
    try:
        response = session.post(
            f"https://{url}/resto/j_spring_security_check?j_username={username}&j_password={password}&_spring_security_remember_me=on&submit=Log+in")
    except Exception as e:
        logger.error(
            "Помилка в запиті авторизації")
        logger.error(e)
        raise Exception()
    if "Failed to log in." in response.text:
        logger.error("Помилка авторизації в iiko")
        raise Exception()
    return response


def generate_reports(session, date, preset_id, headers, port, url):
    try:
        response = session.get(
            f"https://{url}/resto/service/reports/report.jspx?dateFrom={date}&dateTo={date}&presetId={preset_id}", headers=headers)
    except Exception as e:
        logger.error(
            "Помилка в запиті звіту IIKO")
        logger.error(e)
        raise Exception()
    if response.status_code not in [200, 302]:
        logger.error(
            f"Щось пішло не так. Статус код - {response.status_code}")
        raise Exception()
    return response


def questions(restaurants):
    questions = [
        inquirer.Checkbox('restaurant',
                          message="Для якого ресторану скачати звіти?",
                          choices=restaurants,
                          ),
        inquirer.Text('start_date', message="З якої дати скачати звіти?",
                      validate=lambda _, x: re.match(
                          "^[0-9]{1,2}\\.[0-9]{1,2}\\.[0-9]{4}$", x),
                      ),
        inquirer.Text('end_date', message="До якої дати скачати звіти?",
                      validate=lambda _, x: re.match(
                          "^[0-9]{1,2}\\.[0-9]{1,2}\\.[0-9]{4}$", x),
                      ),
    ]
    data = inquirer.prompt(questions)
    return data


def data_processing(data):
    try:
        if data != None:
            if data["restaurant"] == []:
                logger.error("Не обрано жодного ресторану")
                raise Exception()
            logger.info(f"Вибрано такі дані {data}")
            restaurant = data["restaurant"]
            start_date = datetime.strptime(data["start_date"], '%d.%m.%Y')
            end_date = datetime.strptime(data["end_date"], '%d.%m.%Y')
            present = datetime.now()
            if start_date > present or end_date > present:
                logger.error("Одна з введених дат ще не була :)")
                raise Exception()
            return restaurant, start_date, end_date
        else:
            logger.error("Не вибрано нічого")
            raise Exception()
    except Exception as e:
        logger.error(
            "В модулі обробки введених даних сталася помилка:")
        logger.error(e)


def create_excel_report(restaurant, date, xml):
    wb = openpyxl.Workbook()
    ws = wb.active

    element_tree = ET.ElementTree(ET.fromstring(xml))
    root = element_tree.getroot()
    headers = root.find(".//head")
    xml_data = root.findall(".//data")

    ws[var.report_name] = f"Звіт ПН - {restaurant}"
    ws[var.date_cell] = f"Дата: {date}"

    column = 1
    for element in headers:
        ws[f'{get_column_letter(column)}5'] = element.text
        column += 1

    row_num = 1
    for row in xml_data:
        column = 1
        for element in row:
            ws[f"{get_column_letter(column)}{5+row_num}"] = element.text
            column += 1
        row_num += 1

    try:
        set_column_width(ws)
    except Exception as e:
        logger.error(
            "Помилка в модулі встановлення ширини стовпців")
        logger.error(e)
        raise Exception()

    set_formats(ws)
    ex.set_text_format(ws, var.quantity_col)
    try:
        if not os.path.exists(f"./{var.excel_source_path}"):
            os.makedirs(f"./{var.excel_source_path}")
        if not os.path.exists(f"./{var.excel_source_path}/{restaurant}"):
            os.makedirs(f"./{var.excel_source_path}/{restaurant}")
        wb.save(
            f'./{var.excel_source_path}/{restaurant}/{restaurant}_{date}.xlsx')
        logger.info(
            f"\nФайл {restaurant}_{date}.xlsx збережено успішно! \n")
    except Exception as e:
        logger.error(
            "Помилка збереження файлу")
        logger.error(e)
        raise Exception()


def set_column_width(ws):
    dims = {}
    for row in ws.rows:
        for cell in row:
            cell.font = Font(size="9")
            if cell.value:
                dims[cell.column_letter] = max(
                    (dims.get(cell.column_letter, 0), len(str(cell.value))))
    for col, value in dims.items():
        ws.column_dimensions[col].width = value


def set_formats(ws):
    for row in ws.iter_rows(min_row=6, max_row=ws.max_row):
        for cell in row:
            ws[f"H{cell.row}"] = float(ws[f"H{cell.row}"].value)
            ws[f"I{cell.row}"] = float(ws[f"I{cell.row}"].value)




if __name__ == "__main__":
    main()
