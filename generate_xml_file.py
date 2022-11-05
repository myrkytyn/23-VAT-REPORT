from loguru import logger
import variables as var
import xml_generation as xml
from json_info import get_info_xml
import excel as ex
import openpyxl
import os
import json
from datetime import datetime

units_col = "E"


def main():
    time = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    logger.add(f"logs/xml-{time}.log")
    logger.info("### ПРОГРАМА РОЗПОЧАЛА РОБОТУ ###")
    xml_part()
    logger.info("### ПРОГРАМА ЗАКІНЧИЛА РОБОТУ ###")
    try:
        if "ERROR" in open(f"logs/xml-{time}.log", "r", encoding='utf-8').read():
            input(
                "При виконанні програми були помилочки. Перевір чи все гаразд \nНатисни Enter для виходу")
        else:
            input("Мені видається, що все пройшло добре. Натисни Enter для виходу")
    except Exception as e:
        logger.warning("Не вдалося перевірити чи були помилки при виконанні.")
        logger.error(e)
        input("Натисни Enter для виходу")


def xml_part():
    logger.info("### ГЕНЕРАЦІЮ XML ФАЙЛУ РОЗПОЧАТО  ###")
    # Get JSON
    try:
        global config
        config = json.load(open("config.json", "r", encoding="utf-8"))
    except Exception as e:
        logger.error(
            "Схоже, що конфігураційного файлу не існує. Будь ласка. перевір!")
        logger.error(e)
        return
    # For all target excels. Working with XML
    listdir = os.listdir(f"{var.excel_target_path}")
    for dir in listdir:
        print(dir, os.getcwd())
        if os.path.isdir(f"{var.excel_target_path}/{dir}"):
            print(os.getcwd())
            logger.info(f"Створюю податкові ннакладні для {dir}")
            try:
                excel_target_file = ex.list_excels(
                    f"{var.excel_target_path}/{dir}")
            except Exception as e:
                logger.error(
                    f"Помилка, шлях {var.excel_target_path} не знайдено. Поточна директорія - {os.getcwd()}")
                logger.error(e)
                return
            for i in range(0, len(excel_target_file)):
                if excel_target_file[i].endswith(".xlsx"):
                    wb = openpyxl.load_workbook(
                        excel_target_file[i], data_only=True)
                    ws = wb.active
                    iiko_name, hnamesel, tin, hksel, htinsel, hbos, hkbos = get_info_xml(
                        ws, var.restaurant_cell, config)
                    date, year, month = ex.get_date(ws, var.date_cell)
                    hnum = str(ws[f"{var.document_number_cell}"].value)
                    total_without_excise = str(
                        ws[f"{var.sum_without_excise_col}{ws.max_row}"].value)
                    total_without_vat = str(
                        ws[f"{var.sum_without_vat_col}{ws.max_row}"].value)
                    vat_sum = str(
                        ws[f"{var.vat_sum_col}{ws.max_row}"].value)
                    root = xml.generate_root()
                    xml.generate_head(root, tin, period_month=month,
                                      period_year=year, d_fill=date, hnum=hnum)
                    declarbody = xml.generate_body(root=root, hfill=date, hnamesel=hnamesel, hksel=hksel,
                                                   htinsel=htinsel, hnum=hnum, total_without_excise=total_without_excise, total_without_vat=total_without_vat, vat_sum=vat_sum)
                    row_num = 1
                    for row in ws.iter_rows(min_row=6, max_row=ws.max_row-1):
                        cell_row = str(row[0].row)
                        dish = str(ws[f"{var.item_name_col}{cell_row}"].value)
                        if dish == None or dish == "None" or dish == "":
                            dish = str(ws[f"{var.dishes_col}{cell_row}"].value)
                        qnt = str(ws[f"{var.quantity_col}{cell_row}"].value)
                        price = str(ws[f"{var.price_col}{cell_row}"].value)
                        uktzed = str(ws[f"{var.uktzed_col}{cell_row}"].value)
                        row_str = str(row_num)
                        unit_iiko = str(ws[f"{var.units_col}{cell_row}"].value)
                        if unit_iiko == "serv":
                            unit = "порц"
                            unit_code = "3011"
                        elif unit_iiko == "pcs":
                            unit = "пляш"
                            unit_code = "2061"
                        else:
                            unit = "None"
                            unit_code = "None"
                        sum_without_vat = str(
                            ws[f"{var.sum_without_vat_col}{cell_row}"].value)
                        vat_sum = str(round(float(sum_without_vat)*0.2, 3))
                        xml.generate_b_part(declarbody, dish=dish,
                                            row=row_str, qnt=qnt, price=price, uktzed=uktzed, unit=unit, unit_code=unit_code, sum_without_vat=sum_without_vat, vat_sum=vat_sum)
                        row_num += 1
                    xml.generate_ending(declarbody, hbos=hbos, hkbos=hkbos)
                    tree = xml.generate_xml(root)
                    date = ex.get_date(ws, var.date_cell)[0]
                    restaurant = ws[var.restaurant_cell].value
                    cooking_place = ws[var.cooking_place_cell].value
                    rest_dir_name = ex.get_dir_name(
                        restaurant, cooking_place)
                    file_name = f"{var.xml_target_path}{rest_dir_name}/{date}_{iiko_name}.xml"
                    if not os.path.exists(f"../../{var.xml_target_path}{rest_dir_name}"):
                        os.makedirs(
                            f"../../{var.xml_target_path}{rest_dir_name}")
                    try:
                        tree.write(f"../../{file_name}",
                                   encoding="windows-1251", xml_declaration=True)
                    except Exception as e:
                        logger.error(
                            f"Помилка збереження файлу {file_name}, поточна директорія {os.getcwd()}")
                        logger.error(e)
                        return
                    logger.info(f"Файл збережено {file_name}")
            os.chdir("../../")
    logger.info("### XML PART COMPLETED ###")


if __name__ == "__main__":
    main()
