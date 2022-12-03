import json
import openpyxl
import os
from loguru import logger
import json_info as prop
import excel as ex
import get_database as gdb
import variables as var
from datetime import datetime
import inquirer
import argparse


def main():
    time = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    logger.add(f"logs/excel-{time}.log")
    logger.info("### ПРОГРАМА РОЗПОЧАЛА РОБОТУ ###")
    logger.info(f"Версія програми {var.version}")
    use_db, has_args = parse_args()
    if not has_args:
        use_db = questions()
    excel_part(use_db)
    logger.info("### ПРОГРАМА ЗАКІНЧИЛА РОБОТУ ###")
    try:
        if "ERROR" in open(f"./logs/excel-{time}.log", "r", encoding='utf-8').read():
            input(
                "При виконанні програми були помилочки. Перевір чи все гаразд \nНатисни Enter для виходу")
        else:
            input("Мені видається, що все пройшло добре. Натисни Enter для виходу")
    except Exception as e:
        logger.error(e)
        logger.warning("Не вдалося перевірити чи були помилки при виконанні.")
        input("Натисни Enter для виходу")


def parse_args():
    parser = argparse.ArgumentParser()
    parser.add_argument('-db', '--database')

    args = parser.parse_args()
    database = args.database
    if database == None:
        has_args = False
    else:
        has_args = True
        if database == "True":
            database = True
        elif database == "False":
            database = False
    return database, has_args


def questions():
    questions = [
        inquirer.List('use_db',
                      message="Використовувати базу даних?",
                      choices=["Так", "Ні"],
                      ),
    ]
    data = inquirer.prompt(questions)
    global use_db
    if data["use_db"] == "Так":
        use_db = True
    else:
        use_db = False
    return use_db


def excel_part(use_db):
    # Document number input
    #hnum = input("Введи номер, з якого почати нумерувати податкові накладні: ")
    hnum = "1"
    while hnum.isdigit() == False:
        hnum = input(
            "От дідько :( Щось пішло не так\nВведи номер, з якого почати нумерувати податкові накладні: ")
    hnum = int(hnum)
    global config
    config = prop.get_config()
    # For all files in directories
    listdir = os.listdir(f"{var.excel_source_path}")
    for dir in listdir:
        if os.path.isdir(f"{var.excel_source_path}/{dir}"):
            logger.info(f"Створюю Excel ПН для {dir}")
            try:
                excel_source_file = ex.list_excels(
                    f"{var.excel_source_path}/{dir}")
            except Exception as e:
                logger.error(
                    f"Помилка, шлях {var.excel_source_path} не знайдено. Поточна директорія - {os.getcwd()}")
                logger.error(e)
                return
            for file in range(0, len(excel_source_file)):
                if excel_source_file[file].endswith(".xlsx"):
                    # Opening Excel File
                    try:
                        wb = openpyxl.load_workbook(excel_source_file[file])
                        logger.info(
                            f"\n \n Працюю з файлом {excel_source_file[file]} \n \n")
                    except Exception as e:
                        logger.error(
                            f"Помилка відкривання файлу {excel_source_file[file]}")
                        logger.error(e)
                        return
                    ws = wb.active
                    restaurant = ws[var.restaurant_cell].value
                    # Unmerge cells
                    try:
                        ex.unmerge_cells(ws)
                    except Exception as e:
                        logger.error(
                            f"Помилка у модулі роз'єднання клітинок")
                        logger.error(e)
                        continue
                    # Remove delivery rows
                    try:
                        ex.remove_delivery(ws, var.payment_type_col)
                    except Exception as e:
                        logger.error(
                            f"Помилка у модулі видалення доставки")
                        logger.error(e)
                        continue
                    # Clear columns
                    try:
                        ex.clear_cols(ws, var.cost_col, var.percent_col)
                    except Exception as e:
                        logger.error(
                            f"Помилка у модулі очищення непотрібних стовпців")
                        logger.error(e)
                        continue
                    # Remove unused rows
                    try:
                        ex.remove_rows_total(ws, var.cooking_place_col,
                                             var.dishes_group_col)
                    except Exception as e:
                        logger.error(
                            f"Помилка у модулі очищення рядків cум")
                        logger.error(e)
                        continue
                    # remove rows zero price
                    try:
                        ex.remove_rows_zero_price(ws, var.sum_col)
                    except Exception as e:
                        logger.error(
                            f"Помилка у модулі очищення рядків з нульовою ціною")
                        logger.error(e)
                        continue
                    # Sum dish values
                    try:
                        ex.add_values(ws, var.dish_code_col,
                                      var.quantity_col, var.sum_col)
                    except Exception as e:
                        logger.error(
                            f"Помилка у модулі сумування страв")
                        logger.error(e)
                        continue
                    # Calculate without excise from all dishes
                    try:
                        ex.without_excise(
                            ws, var.sum_without_excise_col, var.sum_col)
                    except Exception as e:
                        logger.error(
                            f"Помилка у модулі видалення акцизу")
                        logger.error(e)
                        continue
                    # Get info from JSON file
                    try:
                        iiko_name, non_excise_dishes, non_excise_groups, db_name, groups_to_get_item_names = prop.get_info_xlsx(
                            restaurant, config)
                    except Exception as e:
                        logger.error(
                            f"Помилка у модулі присвєння значень на базі JSON")
                        logger.error(e)
                        continue
                    # Get UKTZED and put it to excel
                    try:
                        if use_db:
                            dish_codes = ex.get_dish_codes(
                                ws, var.dish_code_col)
                            uktzed_codes = gdb.get_uktzed(
                                dish_codes, db_name, config, var.place)
                            if uktzed_codes != "None":
                                ex.uktzed(ws, var.uktzed_col,
                                          uktzed_codes, var.dish_code_col)
                    except Exception as e:
                        logger.error(
                            f"Помилка у модулі обробки кодів УКТЗЕД")
                        logger.error(e)
                        continue
                    # Check and recalculate excise
                    ws[f"{var.dishes_group_col}{ws.max_row+1}"] = "Сума"
                    try:
                        ex.non_excise_dishes_formulas(
                            ws, non_excise_dishes, var.sum_col, var.sum_without_excise_col, var.dishes_col)
                        ex.non_excise_groups_formulas(
                            ws, non_excise_groups, var.sum_col, var.sum_without_excise_col, var.dishes_group_col)
                    except Exception as e:
                        logger.error(
                            f"Помилка у модулі заміни ціни в безакцизних групах та товарах")
                        logger.error(e)
                        continue
                    # Put number of document
                    try:
                        ex.put_hnum(ws, var.document_text_cell,
                                    var.document_number_cell, hnum)
                    except Exception as e:
                        logger.error(
                            f"Помилка в модулі обрахунку порядкового номера документу")
                        logger.error(e)
                        continue
                    # Calculate without VAT
                    try:
                        ex.without_vat(ws, var.sum_without_vat_col,
                                       var.sum_without_excise_col)
                    except Exception as e:
                        logger.error(
                            f"Помилка у модулі розрахунку ціни без ПДВ")
                        logger.error(e)
                        continue
                    # Calculate final price
                    try:
                        ex.price(ws, var.price_col,
                                 var.sum_without_vat_col, var.quantity_col)
                    except Exception as e:
                        logger.error(
                            f"Помилка у модулі розрахунку ціни за одиницю без акцизу та ПДВ")
                        logger.error(e)
                        continue
                    # Calculate VAT sum
                    try:
                        ex.vat_sum(ws, var.sum_without_vat_col,
                                   var.vat_sum_col)
                    except Exception as e:
                        logger.error(
                            f"Помилка у модулі розрахунку суми ПДВ")
                        logger.error(e)
                        continue
                    # Changing Colum width
                    try:
                        ex.change_column_width(
                            ws)
                    except Exception as e:
                        logger.error(
                            f"Помилка у модулі зміни ширини стовпців")
                        logger.error(e)
                        continue
                    # Get Sum of all columns
                    try:
                        ex.get_total(ws, var.sum_without_excise_col,
                                     var.sum_without_vat_col, var.sum_col, var.vat_sum_col)
                    except Exception as e:
                        logger.error(
                            f"Помилка у модулі розрахунку суми для стовпців")
                        logger.error(e)
                        continue
                    # Get item names
                    try:
                        if use_db:
                            ex.item_names(ws, groups_to_get_item_names, var.dish_code_col,
                                          var.dishes_group_col, var.item_name_col, db_name, config, var.place)
                    except Exception as e:
                        logger.error(
                            f"Помилка у модулі визначення назви товару по страві")
                        logger.error(e)
                        continue
                    # Get Date of report
                    try:
                        date = ex.get_date(ws, var.date_cell)[0]
                    except Exception as e:
                        logger.error(
                            f"Помилка у модулі отримання дати за яку створено звіт")
                        logger.error(e)
                        continue
                    file_name = f"{var.excel_target_path}{dir}/{excel_source_file[file]}"
                    ws[var.new_restaurant_cell] = restaurant

                    if not os.path.exists(f"../../{var.excel_target_path}{dir}"):
                        os.makedirs(f"../../{var.excel_target_path}{dir}")
                    try:
                        wb.save(f"../../{file_name}")
                    except Exception as e:
                        logger.error(
                            f"Помилка збереження файлу {file_name}")
                        logger.error(e)
                        continue
                    logger.info(f"Файл збережено {file_name}")
                    hnum += 1
            os.chdir("../../")
    logger.info("### ОБРОБКУ ЗВІТІВ З ФАЙЛІВ ЕКСЕЛЬ ЗАВЕРШЕНО ###")


if __name__ == "__main__":
    main()
