from datetime import datetime, timedelta
import requests
import os
import re
import json_info as prop
from loguru import logger
import json
import inquirer
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
import xml.etree.ElementTree as ET
import variables as var


def main():
    restaurants = []
    headers = {"Accept-Language": "uk,en-US;q=0.7,en;q=0.3"}
    try:
        global config
        config = json.load(open("config.json", "r", encoding="utf-8"))
    except Exception as e:
        logger.error(
            "Схоже, що конфігураційного файлу не існує. Будь ласка. перевір!")
        logger.error(e)
        return
    try:
        username, password = prop.get_info_api(config)
    except Exception as e:
        logger.error(
            "Схоже, що в конфігураційному файлі немає потрібних полів")
        logger.error(e)
        return

    restaurants = get_restaurants(restaurants)
    data = questions(restaurants)
    restaurant, start_date, end_date = data_processing(data)

    #######
    if restaurant == "Урбан Спейс":
        logger.error(
            "Вибач, для Урбан Спейс поки що не працює ;(")
        input("Натисни Enter для виходу")
        return
    #######

    port, preset_id = set_variables(restaurant)
    try:
        delta = end_date - start_date
        logger.info(f"Буде скачано звіти за {delta.days+1} днів")
    except Exception as e:
        logger.error(
            "В модулі підрахунку днів щось пішло не так")
        logger.error(e)
        return
    session = requests.Session()
    try:
        auth(session, port, username, password)
        logger.info("Авторизація пройшла успішно")
    except Exception as e:
        logger.error(
            "В модулі авторизації щось пішло не так")
        logger.error(e)
        return

    for day in range(delta.days+1):
        date = (start_date + timedelta(days=day)).strftime("%d.%m.%Y")
        try: 
            response = generate_reports(
                session, date, preset_id, headers)
        except Exception as e:
            logger.error(
                "В модулі завантаження звітів щось пішло не так")
            logger.error(e)
            return
        try:
            excel_creation(restaurant, date, response.text)
        except Exception as e:
            logger.error(
                "В модулі створення Ексель звіту щось пішло не так")
            logger.error(e)
            return
    input("Натисни Enter для виходу")


def get_restaurants(restaurants):
    for entity in config["legal_entities"]:
        if isinstance(config["legal_entities"][entity]["name"], list):
            restaurants.extend(config["legal_entities"][entity]["name"])
        else:
            restaurants.append(config["legal_entities"][entity]["name"])
    return restaurants


def set_variables(restaurant):
    for entity in config["legal_entities"]:
        if restaurant in config["legal_entities"][entity]["name"]:
            port = config["legal_entities"][entity]["port"]
            if restaurant == "Фабрика-1":
                preset_id = config["legal_entities"][entity]["preset_id"][0]
            elif restaurant == "Фабрика-2":
                preset_id = config["legal_entities"][entity]["preset_id"][1]
            else:
                preset_id = config["legal_entities"][entity]["preset_id"]
    return port, preset_id


def auth(session, port, username, password):
    try:
        response = session.post(
            f"http://iiko.23.ua:{port}/resto/j_spring_security_check?j_username={username}&j_password={password}&_spring_security_remember_me=on&submit=Log+in")
    except Exception as e:
        logger.error(
            "Помилка в запиті авторизації")
        logger.error(e)
        return
    if response.status_code not in [200, 302]:
        logger.error(
            f"Невдала авторизація. Статус код - {response.status_code}")
        return
    return response


def generate_reports(session, date, preset_id, headers):
    try:
        response = session.get(
            f"http://iiko.23.ua:9080/resto/service/reports/report.jspx?dateFrom={date}&dateTo={date}&presetId={preset_id}", headers=headers)
    except Exception as e:
        logger.error(
            "Помилка в запиті звіту IIKO")
        logger.error(e)
        return
    if response.status_code not in [200]:
        logger.error(
            f"Невдала авторизація. Статус код - {response.status_code}")
        return
    return response


def questions(restaurants):
    questions = [
        inquirer.List('restaurant',
                      message="Для якого ресторану скачати звіти?",
                      choices=restaurants,
                      ),
        inquirer.Text('start_date', message="З якої дати скачати звіти?",
                      validate=lambda _, x: re.match(
                          "^[0-9]{1,2}\\.[0-9]{1,2}\\.[0-9]{4}$", x),
                      ),
        inquirer.Text('end_date', message="До якої дати скачати звіти?",
                      validate=lambda _, x: re.match(
                          "^[0-9]{1,2}\\.[0-9]{1,2}\\.[0-9]{4}$", x),
                      ),
    ]
    data = inquirer.prompt(questions)
    return data


def data_processing(data):
    if data != None:
        logger.info(f"Вибрано такі дані {data}")
        restaurant = data["restaurant"]
        start_date = datetime.strptime(data["start_date"], '%d.%m.%Y')
        end_date = datetime.strptime(data["end_date"], '%d.%m.%Y')
        return restaurant, start_date, end_date
    else:
        logger.warning(f"Не вибрано нічого")
        return


def excel_creation(restaurant, date, xml):
    wb = openpyxl.Workbook()
    ws = wb.active

    element_tree = ET.ElementTree(ET.fromstring(xml))
    root = element_tree.getroot()
    headers = root.find(".//head")
    xml_data = root.findall(".//data")

    ws[var.report_name] = f"Звіт ПН - {restaurant}"
    ws[var.date_cell] = f"Дата: {date}"

    column = 1
    for element in headers:
        ws[f'{get_column_letter(column)}5'] = element.text
        column += 1

    row_num = 1
    for row in xml_data:
        column = 1
        for element in row:
            ws[f"{get_column_letter(column)}{5+row_num}"] = element.text
            column += 1
        row_num += 1

    try:
        set_column_width(ws)
    except Exception as e:
        logger.error(
            "Помилка в модулі встановлення ширини стовпців")
        logger.error(e)
        return
    try:
        if not os.path.exists(f"./{var.excel_source_path}"):
            os.makedirs(f"./{var.excel_source_path}")
        if not os.path.exists(f"./{var.excel_source_path}/{restaurant}"):
            os.makedirs(f"./{var.excel_source_path}/{restaurant}")
        wb.save(
            f'./{var.excel_source_path}/{restaurant}/{restaurant}_{date}.xlsx')
        logger.info(
            f"\nФайл {restaurant}_{date}.xlsx збережено успішно! \n")
    except Exception as e:
        logger.error(
            "Помилка збереження файлу")
        logger.error(e)
        return


def set_column_width(ws):
    dims = {}
    for row in ws.rows:
        for cell in row:
            cell.font = Font(size="9")
            if cell.value:
                dims[cell.column_letter] = max(
                    (dims.get(cell.column_letter, 0), len(str(cell.value))))
    for col, value in dims.items():
        ws.column_dimensions[col].width = value


if __name__ == "__main__":
    main()
