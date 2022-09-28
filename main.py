from genericpath import exists
import json
from tokenize import String
from xml.etree.ElementTree import tostring
import openpyxl
import os
from loguru import logger
from xml_generation import generate_xml, generate_root, generate_head, generate_body, generate_b_part, generate_ending
from json_info import get_info_xlsx, get_info_xml
import excel as ex
import get_database as gdb


excel_source_path = "./iiko_reports/"
excel_target_path = "./excel_files_generated/"
xml_target_path = "./xml_files_generated/"

date_cell = "A3"
restaurant_cell = "A6"
cooking_place_col = "B"
cooking_place_cell = "B6"
dishes_group_col = "C"
dishes_col = "D"
units_col = "E"
dish_code_col = "F"
quantity_col = "G"
sum_col = "H"
cost_col = "I"
percent_col = "J"
sum_without_excise_col = "L"
sum_without_vat_col = "M"
price_col = "N"
uktzed_col = "O"
document_text_cell = "A4"
document_number_cell = "B4"


def main():
    logger.add("logs/{time}.log")
    logger.info("### ПРОГРАМА РОЗПОЧАЛА РОБОТУ ###")
    excel_part()
    #input("Press Enter to continue...")
    # xml_part()
    logger.info("### ПРОГРАМА ЗАКІНЧИЛА РОБОТУ ###")
    input("Мені видається, що все пройшло добре. Натисни Enter для виходу")


def excel_part():
    # Work with config.json
    hnum = input("Введи номер, з якого почати нумерувати податкові накладні: ")
    while hnum.isdigit() == False:
        hnum = input(
            "От дідько :( Щось пішло не так\nВведи номер, з якого почати нумерувати податкові накладні: ")
    hnum = int(hnum)
    try:
        global config
        config = json.load(open("config.json", "r", encoding="utf-8"))
    except Exception as e:
        logger.error(
            "Схоже, що конфігураційного файлу не існує. Будь ласка. перевір!")
        logger.error(e)
        return
    # Get all files from directory
    logger.info("### ОБРОБКУ ЗВІТІВ З ФАЙЛІВ ЕКСЕЛЬ РОЗПОЧАТО ###")
    try:
        excel_source_file = ex.list_excels(excel_source_path)
    except Exception as e:
        logger.error(
            f"Помилка, шлях {excel_source_path} не знайдено. Поточна директорія - {os.getcwd()}")
        logger.error(e)
        return
    for i in range(0, len(excel_source_file)):
        if excel_source_file[i].endswith(".xlsx"):
            try:
                wb = openpyxl.load_workbook(excel_source_file[i])
            except Exception as e:
                logger.error(
                    f"Помилка відкривання файлу {excel_source_file[i]}")
                logger.error(e)
                return
            ws = wb.active
            # 1 - unmerge cells
            ex.unmerge_cells(ws)
            # clear columns
            ex.clear_cols(ws, cost_col, percent_col)
            # 2 - remove unused rows
            ex.remove_rows_total(ws, cooking_place_col, dishes_group_col)
            ex.remove_rows_zero_price(ws, sum_col)
            # 3 - calculate without excise fromm all dishes
            ex.without_excise(ws, sum_without_excise_col, sum_col)
            # 4 -
            iiko_name, non_excise_dishes, non_excise_groups, db_name = get_info_xlsx(
                ws, restaurant_cell, config)
            # UKTZED
            dish_codes = ex.get_dish_codes(ws, dish_code_col)
            uktzed_codes = gdb.get_uktzed(dish_codes, db_name)
            if uktzed_codes != "None":
                ex.uktzed(ws, uktzed_col, uktzed_codes, dishes_col)
            # 5 - check and recalculate excise
            ex.non_excise_dishes_formulas(
                ws, non_excise_dishes, sum_col, sum_without_excise_col)
            ex.non_excise_groups_formulas(
                ws, non_excise_groups, sum_col, sum_without_excise_col)
            # 6 - put number of document
            ex.put_hnum(ws, document_text_cell, document_number_cell, hnum)
            # 7 - calculate without VAT
            ex.without_vat(ws, sum_without_vat_col, sum_without_excise_col)
            # 8 - calculate final price
            ex.price(ws, price_col, sum_without_vat_col, quantity_col)
            date = ex.get_date(ws, date_cell)[0]
            rest_dir_name = ex.get_dir_name(
                ws, restaurant_cell, cooking_place_cell)
            file_name = f"{excel_target_path}{rest_dir_name}/{date}_{iiko_name}.xlsx"
            ex.change_column_width(
                ws, sum_without_excise_col, sum_without_vat_col, price_col, uktzed_col)
            ex.get_total(ws, sum_without_excise_col,
                         sum_without_vat_col, sum_col)
            if not os.path.exists(f"../{excel_target_path}{rest_dir_name}"):
                os.makedirs(f"../{excel_target_path}{rest_dir_name}")
            try:
                wb.save(f"../{file_name}")
            except Exception as e:
                logger.error(
                    f"Помилка збереження файлу {file_name}")
                logger.error(e)
                return
            logger.info(f"Файл збережено {file_name}")
            hnum += 1
    logger.info("### ОБРОБКУ ЗВІТІВ З ФАЙЛІВ ЕКСЕЛЬ ЗАВЕРШЕНО ###")


def xml_part():
    logger.info("### ГЕНЕРАЦІЮ XML ФАЙЛУ РОЗПОЧАТО  ###")
    os.chdir("..")
    # For all target excels. Working with XML
    try:
        excel_target_file = ex.list_excels(excel_target_path)
    except Exception as e:
        logger.error(
            f"Помилка, щлях {excel_target_path} не знайдено. Поточна директорія - {os.getcwd()}")
        logger.error(e)
        return
    for i in range(0, len(excel_target_file)):
        if excel_target_file[i].endswith(".xlsx"):
            wb = openpyxl.load_workbook(excel_target_file[i], data_only=True)
            ws = wb.active
            iiko_name, hnamesel, tin, hksel, htinsel, hbos, hkbos = get_info_xml(
                ws, restaurant_cell, config)
            date, year, month = ex.get_date(ws, date_cell)
            root = generate_root()
            generate_head(root, tin, period_month=month,
                          period_year=year, d_fill=date)
            declarbody = generate_body(root=root, hfill=date, hnamesel=hnamesel, hksel=hksel,
                                       htinsel=htinsel)
            row_num = 1
            for row in ws.iter_rows(min_row=6, max_row=ws.max_row):
                cell_row = str(row[0].row)
                dish = str(ws[f"{dishes_col}{cell_row}"].value)
                qnt = str(ws[f"{quantity_col}{cell_row}"].value)
                price = str(ws[f"{price_col}{cell_row}"].value)
                row_str = str(row_num)
                generate_b_part(declarbody, dish=dish,
                                row=row_str, qnt=qnt, price=price)
                row_num += 1
            generate_ending(declarbody, hbos=hbos, hkbos=hkbos)
            tree = generate_xml(root)
            if not os.path.exists(f"../{xml_target_path}"):
                os.makedirs(f"../{xml_target_path}")
            try:
                tree.write(f"../{xml_target_path}{date}_{iiko_name}.xml",
                           encoding="windows-1251", xml_declaration=True)
            except Exception as e:
                logger.error(
                    f"Помилка збереження файлу {xml_target_path}{date}_{iiko_name}.xml")
                logger.error(e)
                return
            logger.info(
                f"Файл збережено {xml_target_path}{date}_{iiko_name}.xml")
    logger.info("### XML PART COMPLETED ###")


if __name__ == "__main__":
    main()
