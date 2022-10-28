from loguru import logger
import os
import re
import openpyxl
import variables as var
from openpyxl.styles import Font
from openpyxl.utils import column_index_from_string
import get_database as db

def list_excels(excel_path):
    os.chdir(excel_path)
    excel_files = os.listdir('.')
    return excel_files


def unmerge_cells(ws):
    for merge in list(ws.merged_cells):
        ws.unmerge_cells(range_string=str(merge))


def remove_delivery(ws, payment_type_col):
    for cell in ws[payment_type_col]:
        if cell.value == "Доставка":
            logger.info("Доставку знайдено. Видалення рядків запущено")
            i = 1
            while ws[f"{payment_type_col}{cell.row+i}"].value == None:
                i += 1
            ws.delete_rows(cell.row, cell.row+i)
            logger.info(
                f"Видалено рядки з доставкою - {cell.row}:{cell.row+i}")
    for cell in ws[payment_type_col]:
        if cell.value is not None:
            if "Total" in cell.value or "всего" in cell.value:
                ws.delete_rows(cell.row)
    ws.delete_cols(
        openpyxl.utils.cell.column_index_from_string(payment_type_col))
    logger.info("Стовпець з типом оплат видалено")


def add_values(ws, dish_code_col, quantity_col, sum_col):
    rows_to_remove = []
    for row in ws.iter_rows(min_col=4, max_col=4, min_row=6, max_row=ws.max_row):
        for cell in row:
            for next_row in ws.iter_rows(min_col=4, max_col=4, min_row=cell.row+1, max_row=ws.max_row):
                for cell1 in next_row:
                    if ws[f"{dish_code_col}{cell.row}"].value == ws[f"{dish_code_col}{cell1.row}"].value:
                        ws[f"{quantity_col}{cell.row}"] = float(ws[f"{quantity_col}{cell.row}"].value) + \
                            float(ws[f"{quantity_col}{cell1.row}"].value)
                        ws[f"{sum_col}{cell.row}"] = float(ws[f"{sum_col}{cell.row}"].value) + \
                            float(ws[f"{sum_col}{cell1.row}"].value)
                        rows_to_remove.append(cell1.row)
    rows_to_remove.sort()
    rows_to_remove = list(dict.fromkeys(rows_to_remove))
    i = 0
    for row_to_remove in rows_to_remove:
        ws.delete_rows(row_to_remove-i)
        i += 1


def remove_rows_total(ws, cooking_place_col, dishes_group_col):
    # Removing rows with Total in column - Місце приготування
    for cell in ws[cooking_place_col]:
        if cell.value is not None:
            if "Total" in cell.value or "всего" in cell.value:
                ws.delete_rows(cell.row)
    # Removing rows with Total in column - Група страви
    for cell in ws[dishes_group_col]:
        if cell.value is not None:
            if "Total" in cell.value or "всего" in cell.value:
                ws.delete_rows(cell.row)


def remove_rows_zero_price(ws, sum_col):
    num_col=column_index_from_string(sum_col)
    rows_to_remove=[]
    for row in ws.iter_rows(min_col=num_col, max_col=num_col, min_row=6, max_row=ws.max_row):
        for cell in row:
            if cell.value == 0 or cell.value == "0":
                if (ws[f"{var.dishes_group_col}{cell.row}"].value != None and ws[f"{var.dishes_group_col}{cell.row+1}"].value == None and cell.row+1 < ws.max_row):
                    ws[f"{var.dishes_group_col}{cell.row+1}"] = ws[f"{var.dishes_group_col}{cell.row}"].value
                rows_to_remove.append(cell.row)
    rows_to_remove.sort()
    rows_to_remove = list(dict.fromkeys(rows_to_remove))
    i = 0
    for row_to_remove in rows_to_remove:
        ws.delete_rows(row_to_remove-i)
        i += 1


def without_excise(ws, sum_without_excise_col, sum_col):
    ws[f"{sum_without_excise_col}5"] = "Сума без акцизу"
    for row in ws.iter_rows(min_col=4, max_col=4, min_row=6, max_row=ws.max_row-1):
        for cell in row:
            # Find price without excise
            sum_without_excise_fl = round(
                float(ws[f"{sum_col}{cell.row}"].value)/105*100, 3)
            ws[f"{sum_without_excise_col}{cell.row}"] = sum_without_excise_fl


def without_vat(ws, sum_without_vat_col, sum_without_excise_col):
    ws[f"{sum_without_vat_col}5"] = "Сума без ПДВ"
    for row in ws.iter_rows(min_col=4, max_col=4, min_row=6, max_row=ws.max_row-2):
        for cell in row:
            # Find price without VAT
            sum_without_vat_fl = round(
                (float(ws[f"{sum_without_excise_col}{cell.row}"].value)/6*5), 3)
            ws[f"{sum_without_vat_col}{cell.row}"] = sum_without_vat_fl


def price(ws, price_col, sum_without_vat_col, quantity_col):
    ws[f"{price_col}5"] = "Ціна без ПДВ"
    for row in ws.iter_rows(min_col=4, max_col=4, min_row=6, max_row=ws.max_row-2):
        for cell in row:
            # Find price for 1 product
            ws[f"{price_col}{cell.row}"] = round((float(
                ws[f"{sum_without_vat_col}{cell.row}"].value)/float(ws[f"{quantity_col}{cell.row}"].value)), 3)


def uktzed(ws, uktzed_col, uktzed_codes, dish_code_col):
    ws[f"{uktzed_col}5"] = "Код УКТЗЕД"
    for row in ws.iter_rows(min_col=4, max_col=4, min_row=6, max_row=ws.max_row):
        for cell in row:
            for r in uktzed_codes:
                if ws[f"{dish_code_col}{cell.row}"].value == r[0]:
                    ws[f"{uktzed_col}{cell.row}"] = r[1]


def non_excise_dishes_formulas(ws, non_excise_dishes, sum_col, sum_without_excise_col, dishes_col):
    counter = 0
    for cell in ws[dishes_col]:
        if cell.value in non_excise_dishes:
            ws[f"{sum_without_excise_col}{cell.row}"] = ws[f"{sum_col}{cell.row}"].value
            counter += 1
    logger.info(f"Акциз видалено з {counter} страв")


def non_excise_groups_formulas(ws, non_excise_groups, sum_col, sum_without_excise_col, dishes_group_col):
    counter_groups = 0
    counter_dishes = 0
    for cell in ws[dishes_group_col]:
        if cell.value in non_excise_groups:
            ws[f'{sum_without_excise_col}{cell.row}'] = ws[f"{sum_col}{cell.row}"].value
            counter_dishes += 1
            i = 1
            while ws[f"C{cell.row+i}"].value == None:
                ws[f'{sum_without_excise_col}{cell.row+i}'] = ws[f"{sum_col}{cell.row+i}"].value
                counter_dishes += 1
                i += 1
            counter_groups += 1
    logger.info(
        f"Акциз видалено з {counter_dishes} страв в {counter_groups} групах")


def get_date(ws, date_cell):
    # date to format ddmmyyyy
    date = re.sub("[^0-9]", "", ws[date_cell].value)
    year = date[-4:]
    if date[2] == "0":
        month = date[3]
    else:
        month = date[2]+date[3]
    return (date, year, month)


def get_dir_name(restaurant, cooking_place):
    dir_suffix = cooking_place
    dir_suffix = dir_suffix[dir_suffix.find('ф'):]
    dir_name = f"{restaurant}-{dir_suffix}"
    return dir_name


def put_hnum(ws, text_cell, num_cell, hnum):
    ws[text_cell] = "Номер документу"
    ws[num_cell] = hnum


def change_column_width(ws):
    dims = {}
    for row in ws.rows:
        for cell in row:
            cell.font = Font(size="9")
            if cell.value:
                dims[cell.column_letter] = max(
                    (dims.get(cell.column_letter, 0), len(str(cell.value))))
    for col, value in dims.items():
        ws.column_dimensions[col].width = value


def get_total(ws, sum_without_excise_col, sum_without_vat_col, sum_col):
    total_sum = 0
    total_without_excise = 0
    total_without_vat = 0
    max_row = ws.max_row
    for row in ws.iter_rows(min_col=4, max_col=4, min_row=6, max_row=ws.max_row-2):
        for cell in row:
            total_without_vat += float(ws[f"{sum_without_vat_col}{cell.row}"].value)
            total_without_excise += float(ws[f"{sum_without_excise_col}{cell.row}"].value)
            total_sum += float(ws[f"{sum_col}{cell.row}"].value)
    ws[f"{sum_without_excise_col}{max_row}"] = round(total_without_excise, 2)
    ws[f"{sum_without_vat_col}{max_row}"] = round(total_without_vat, 2)
    ws[f"{sum_col}{max_row}"] = round(total_sum, 2)


def get_dish_codes(ws, dish_code_col):
    dish_codes = ""
    dish_code = ""
    for row in ws.iter_rows(min_col=4, max_col=4, min_row=6, max_row=ws.max_row):
        for cell in row:
            dish_code = ws[f"{dish_code_col}{cell.row}"].value
            dish_codes += f"'{dish_code}',"
    dish_codes = dish_codes[:-1]
    return dish_codes


def clear_cols(ws, col_to_remove1, col_to_remove2):
    for row in ws[f'{col_to_remove1}1:{col_to_remove1}{ws.max_row}']:
        for cell in row:
            cell.value = None

    for row in ws[f'{col_to_remove2}1:{col_to_remove2}{ws.max_row}']:
        for cell in row:
            cell.value = None

def item_names(ws, groups_to_get_item_names, dish_code_col, dishes_group_col, item_name_col, DATABASE, config, place):
    ws[f"{item_name_col}5"] = "Назва товару"
    for cell in ws[dishes_group_col]:
        if cell.value in groups_to_get_item_names:
            num = ws[f"{dish_code_col}{cell.row}"].value
            ws[f'{item_name_col}{cell.row}'] = db.get_item_name(num, DATABASE, config, place)
            i = 1
            while ws[f"C{cell.row+i}"].value == None:
                num = ws[f"{dish_code_col}{cell.row+i}"].value
                ws[f'{item_name_col}{cell.row+i}'] = db.get_item_name(num, DATABASE, config, place)
                i += 1
    logger.info(
        f"Знайдено товари по стравах")