from __future__ import print_function
import os
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
import variables as var
import excel as ex
import openpyxl
from loguru import logger
import re
import json_info as prop
from prettytable import PrettyTable, ALL


SCOPES = ['https://www.googleapis.com/auth/spreadsheets.readonly']
R = "\033[0;31;40m"
N = "\033[0m"

def main():
    config = prop.get_config()
    sheet_id = config.get("test_module").get("sheet_id")
    listdir = sorted(os.listdir(f"{var.excel_target_path}"))
    for dir in listdir:
        table = PrettyTable()
        table.field_names = ["Дата", "Google", "Excel"]
        table.hrules = ALL
        if os.path.isdir(f"{var.excel_target_path}/{dir}"):
            excel_files = ex.list_excels(
                f"{var.excel_target_path}{dir}")
            sheet = get_sheet_name(config, dir)
            table.add_row([dir, "", ""])
            if dir == "Урбан Спейс":
                col_date = "A"
                col_sum = "E"
            else:
                col_date = "D"
                col_sum = "H"
            for i in range(0, len(excel_files)):
               if excel_files[i].endswith(".xlsx"):
                   wb = openpyxl.load_workbook(
                       f"{var.excel_target_path}{dir}/{excel_files[i]}", data_only=True)
                   ws = wb.active
                   date = re.sub("[^0-9.]", "", ws[var.date_cell].value)
                   full_date = date
                   try:
                       result = get_values(sheet_id, f"{sheet}!{col_date}:{col_date}")
                   except Exception as e:
                       logger.error(
                           f"Проблеми з конектом до Google Sheets {os.getcwd()}")
                       logger.error(e)
                   try:
                       index = result["values"].index([full_date])+1
                   except Exception as e:
                       logger.error(
                           f"Щось з датою")
                       logger.error(e)
                       continue
                   sum_excel = float(ws[f"{var.sum_col}{ws.max_row}"].value)
                   #sum_vat_excel = float(
                   #    ws[f"{var.vat_sum_col}{ws.max_row}"].value)
                   sum_gsheet = float(get_values(sheet_id, f"{sheet}!{col_sum}{index}")[
                                      "values"][0][0].replace(',', '.'))
                   #sum_vat_gsheet = float(get_values(sheet_id, f"{sheet}!P{index}")[
                   #                       "values"][0][0].replace(',', '.'))
                   if sum_excel == sum_gsheet:
                       table.add_row([full_date, sum_gsheet, sum_excel])
                   else:
                       table.add_row([full_date, R+str(sum_gsheet)+N, R+str(sum_excel)+N])
                   #if sum_vat_excel == sum_vat_gsheet:
                   #    logger.info(
                   #        f"Сума ПДВ співпадає! {sum_vat_excel} = {sum_vat_gsheet}")
                   #else:
                   #    logger.warning(
                   #        f"Сума ПДВ в ексель - {sum_vat_excel}, в гугл таблицях - {sum_vat_gsheet}")
        print(table)
            

def get_values(spreadsheet_id, range_name):
    creds=None
    # The file token.json stores the user's access and refresh tokens, and is
    # created automatically when the authorization flow completes for the first
    # time.
    if os.path.exists('token.json'):
        creds=Credentials.from_authorized_user_file('token.json', SCOPES)
    # If there are no (valid) credentials available, let the user log in.
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow=InstalledAppFlow.from_client_secrets_file(
                'credentials.json', SCOPES)
            creds=flow.run_local_server(port=0)
        # Save the credentials for the next run
        with open('token.json', 'w') as token:
            token.write(creds.to_json())
    try:
        service=build('sheets', 'v4', credentials=creds)

        result=service.spreadsheets().values().get(
            spreadsheetId=spreadsheet_id, range=range_name).execute()
        rows=result.get('values', [])
        return result
    except HttpError as error:
        logger.error(f"An error occurred: {error}")
        return error


def get_sheet_name(config, restaurant):
    for entity in config["legal_entities"]:
        if restaurant in config["legal_entities"][entity]["name"]:
            if isinstance(config["legal_entities"][entity]["name"], list):
                position=config["legal_entities"][entity]["name"].index(
                    restaurant)
                sheet=config["legal_entities"][entity]["sheet"][position]
            else:
                sheet=config["legal_entities"][entity]["sheet"]
    return sheet


if __name__ == '__main__':
    main()
